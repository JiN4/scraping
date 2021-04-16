# -*- coding: utf-8 -*-
from datetime import datetime
import openpyxl as px
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
 
url = 'https://www.ryokolink.com/travel/hotelsogo_2021open_hyogo.htm';
 
r    = requests.get(url)
r.encoding = r.apparent_encoding
soup = BeautifulSoup(r.text, 'lxml')


def Pagecrawling(soup, item_list):
  # place = soup.find_all('li', class_='li_shukuhaku_submenu_now')
  ryokans = soup.find_all('dl')
 
  for ryokan in ryokans:
    name = ryokan.find('dt', class_='').string
    
    item_list.append(name)

  return item_list
 

def Write_excel(item_list):
  wb = px.Workbook()
  ws = wb.active
 
  fill = PatternFill(patternType='solid', fgColor='e0e0e0', bgColor='e0e0e0')
 
  headers = ['名前']
  for i, header in enumerate(headers):
    ws.cell(row=1, column=1+i, value=headers[i])
    ws.cell(row=1, column=1+i).fill = fill
 
    for y, row in enumerate(item_list):
      ws.cell(row= y+2, column= 1, value=item_list[y])

    now = datetime.now()
    data = now.strftime('%Y-%m-%d')
 
    filename = 'ryokolink' + data + '.xlsx'
    wb.save(filename)

item_list = []
item_list = Pagecrawling(soup, item_list)
Write_excel(item_list)
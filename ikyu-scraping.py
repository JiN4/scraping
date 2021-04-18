# -*- coding: utf-8 -*-
from datetime import datetime
import openpyxl as px
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
import re

def Pagecrawling(soup, item_list):
  ryokans = soup.find_all('div', class_='acc_list_each')
 
  for ryokan in ryokans:
    names   = ryokan.find_all('h2',   class_='acmName')
    places  = ryokan.find_all('div',  class_='acmArea')
    reviews = ryokan.find_all('span', class_='acmEvaluationCnt')
    prices  = ryokan.find_all('span', class_='roomLowestAmount')

    for (name, place, review, price) in zip(names, places, reviews, prices):
      nameEx    = name.get_text()
      placeEx   = place.get_text().strip()
      priceEx   = price.get_text()
      reviewStr = review.get_text().replace(' ', '').replace('\n' , '')
      reviewEx = re.search(r'^(.+)$', reviewStr).group().replace('\r' , '')

      item_list.append([nameEx, placeEx, priceEx, reviewEx])

  return item_list
 
def Write_excel(item_list):
  wb = px.Workbook()
  ws = wb.active
 
  fill = PatternFill(patternType='solid', fgColor='e0e0e0', bgColor='e0e0e0')
 
  headers = ['名前', '地域', '最低金額(2名)', '評価']
  for i, header in enumerate(headers):
    ws.cell(row=1, column=1+i, value=headers[i])
    ws.cell(row=1, column=1+i).fill = fill
 
    for y, row in enumerate(item_list):
      for x, cell in enumerate(row):
        ws.cell(row= y+2, column= x+1, value=item_list[y][x])

    now = datetime.now()
    data = now.strftime('%Y-%m-%d')
 
    filename = 'Ikyu' + data + '.xlsx'
    wb.save(filename)

item_list = []
urls      = ['https://www.ikyu.com/hokuriku/220000/p1/?hoi=2&gotoFlag=0',
             'https://www.ikyu.com/hokuriku/220000/p2/?hoi=2&gotoFlag=0',
             'https://www.ikyu.com/hokuriku/220000/p3/?hoi=2&gotoFlag=0',
             'https://www.ikyu.com/hokuriku/220000/p4/?hoi=2&gotoFlag=0',
             'https://www.ikyu.com/hokuriku/220000/p5/?hoi=2&gotoFlag=0',
             'https://www.ikyu.com/hokuriku/220000/p6/?hoi=2&gotoFlag=0',
             'https://www.ikyu.com/hokuriku/220000/p7/?hoi=2&gotoFlag=0']

for url in urls:
  r    = requests.get(url)
  soup = BeautifulSoup(r.text, 'lxml')
  item_list = Pagecrawling(soup, item_list)

Write_excel(item_list)